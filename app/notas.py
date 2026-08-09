import os
import uuid
from datetime import datetime
from decimal import Decimal, InvalidOperation

from flask import (
    Blueprint, render_template, request, redirect, url_for, flash,
    send_from_directory, current_app, abort, make_response
)

from flask_login import login_required, current_user
from sqlalchemy import func

from . import db
from .models import Nota, User, Categoria, LogAuditoria

notas_bp = Blueprint("notas", __name__)


def registrar_log(acao, detalhes=None):
    """Registra uma acao no log de auditoria do sistema."""
    try:
        ip = request.headers.get("X-Forwarded-For", request.remote_addr)
        if ip and "," in ip:
            ip = ip.split(",")[0].strip()
        log = LogAuditoria(
            usuario_id=current_user.id if current_user.is_authenticated else None,
            usuario_nome=current_user.nome if current_user.is_authenticated else "Sistema",
            acao=acao,
            detalhes=detalhes,
            ip_address=ip
        )
        db.session.add(log)
        db.session.commit()
    except Exception:
        db.session.rollback()



def arquivo_permitido(nome_arquivo):
    ext = nome_arquivo.rsplit(".", 1)[-1].lower() if "." in nome_arquivo else ""
    return ext in current_app.config["ALLOWED_EXTENSIONS"]


def salvar_arquivo(arquivo):
    ext = arquivo.filename.rsplit(".", 1)[-1].lower()
    nome_salvo = f"{uuid.uuid4().hex}.{ext}"
    caminho = os.path.join(current_app.config["UPLOAD_FOLDER"], nome_salvo)
    arquivo.save(caminho)
    return nome_salvo


def parse_data(valor_str, padrao=None):
    try:
        return datetime.strptime(valor_str, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return padrao


def parse_valor(valor_str):
    try:
        valor_str = (valor_str or "0").replace(".", "").replace(",", ".") \
            if "," in (valor_str or "") else (valor_str or "0")
        return Decimal(valor_str)
    except InvalidOperation:
        return Decimal("0")


@notas_bp.route("/")
@login_required
def dashboard():
    query = Nota.query

    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")
    venc_inicio = request.args.get("venc_inicio", "")
    venc_fim = request.args.get("venc_fim", "")
    tipo = request.args.get("tipo", "")
    categoria = request.args.get("categoria", "")
    tipo_documento = request.args.get("tipo_documento", "")
    status_filtro = request.args.get("status", "")
    cliente = request.args.get("cliente", "").strip()
    cadastrado_por = request.args.get("cadastrado_por", "")
    somente_atrasadas = request.args.get("somente_atrasadas", "")
    busca = request.args.get("busca", "").strip()

    hoje = datetime.utcnow().date()

    if data_inicio:
        d = parse_data(data_inicio)
        if d:
            query = query.filter(Nota.data_emissao >= d)
    if data_fim:
        d = parse_data(data_fim)
        if d:
            query = query.filter(Nota.data_emissao <= d)
    if venc_inicio:
        d = parse_data(venc_inicio)
        if d:
            query = query.filter(Nota.data_vencimento >= d)
    if venc_fim:
        d = parse_data(venc_fim)
        if d:
            query = query.filter(Nota.data_vencimento <= d)
    if tipo in ("entrada", "saida"):
        query = query.filter(Nota.tipo == tipo)
    if categoria:
        query = query.filter(Nota.categoria == categoria)
    if tipo_documento:
        query = query.filter(Nota.tipo_documento == tipo_documento)
    if status_filtro:
        query = query.filter(Nota.status == status_filtro)
    if cadastrado_por and cadastrado_por.isdigit():
        query = query.filter(Nota.enviado_por_id == int(cadastrado_por))
    if somente_atrasadas == "1":
        query = query.filter(Nota.status == "Pendente", Nota.data_vencimento < hoje)
    if cliente:
        query = query.filter(Nota.cliente_fornecedor.ilike(f"%{cliente}%"))
    if busca:
        like = f"%{busca}%"
        query = query.filter(
            db.or_(
                Nota.cliente_fornecedor.ilike(like),
                Nota.numero_nota.ilike(like),
                Nota.descricao.ilike(like),
            )
        )

    notas = query.order_by(Nota.data_vencimento.asc().nullslast(), Nota.data_emissao.desc(), Nota.id.desc()).all()

    # Cálculos para resumo financeiro detalhado
    total_entradas = sum((n.valor for n in notas if n.tipo == "entrada" and n.status == "Recebido"), Decimal("0"))
    total_saidas = sum((n.valor for n in notas if n.tipo == "saida" and n.status == "Pago"), Decimal("0"))
    
    total_a_receber = sum((n.valor for n in notas if n.tipo == "entrada" and n.status == "Pendente"), Decimal("0"))
    total_a_pagar = sum((n.valor for n in notas if n.tipo == "saida" and n.status == "Pendente"), Decimal("0"))
    
    total_atrasadas_valor = sum((n.valor for n in notas if n.is_atrasada), Decimal("0"))
    qtd_atrasadas = sum(1 for n in notas if n.is_atrasada)

    # Lista de clientes/fornecedores distintos para o dropdown
    clientes = [
        r[0] for r in
        db.session.query(Nota.cliente_fornecedor)
        .filter(Nota.cliente_fornecedor.isnot(None))
        .filter(Nota.cliente_fornecedor != "")
        .distinct()
        .order_by(Nota.cliente_fornecedor)
        .all()
    ]

    # Lista de usuários cadastrado por para o dropdown
    usuarios = User.query.order_by(User.nome).all()

    # Agrupamento para gráfico por Categoria
    categorias_dict = {}
    for n in notas:
        cat = n.categoria or "Outro"
        categorias_dict[cat] = categorias_dict.get(cat, Decimal("0")) + n.valor

    chart_labels = list(categorias_dict.keys())
    chart_values = [float(v) for v in categorias_dict.values()]

    return render_template(
        "dashboard.html",
        notas=notas,
        total_entradas=total_entradas,
        total_saidas=total_saidas,
        total_a_receber=total_a_receber,
        total_a_pagar=total_a_pagar,
        total_atrasadas_valor=total_atrasadas_valor,
        qtd_atrasadas=qtd_atrasadas,
        saldo=total_entradas - total_saidas,
        filtros=request.args,
        clientes=clientes,
        usuarios=usuarios,
        hoje=hoje,
        chart_labels=chart_labels,
        chart_values=chart_values,
    )




@notas_bp.route("/notas/nova", methods=["GET", "POST"])
@login_required
def upload():
    if request.method == "POST":
        arquivo = request.files.get("arquivo")

        if not arquivo or arquivo.filename == "":
            flash("Selecione um arquivo (PDF ou foto da nota).", "danger")
            return redirect(url_for("notas.upload"))

        if not arquivo_permitido(arquivo.filename):
            flash("Formato não permitido. Envie PDF, JPG, PNG ou WEBP.", "danger")
            return redirect(url_for("notas.upload"))

        data_emissao = parse_data(request.form.get("data_emissao"), datetime.utcnow().date())
        data_vencimento = parse_data(request.form.get("data_vencimento"), data_emissao)
        data_pagamento = parse_data(request.form.get("data_pagamento"), None)
        status = request.form.get("status", "Pendente")

        comprovante = request.files.get("comprovante")
        comprovante_salvo = None
        comprovante_orig = None
        if comprovante and comprovante.filename:
            if arquivo_permitido(comprovante.filename):
                comprovante_salvo = salvar_arquivo(comprovante)
                comprovante_orig = comprovante.filename

        valor_pago_val = parse_valor(request.form.get("valor_pago"))
        valor_val = parse_valor(request.form.get("valor"))
        if status == "Pendente" and valor_pago_val > 0 and valor_pago_val < valor_val:
            status = "Parcial"

        nota = Nota(
            numero_nota=request.form.get("numero_nota", "").strip(),
            tipo=request.form.get("tipo", "entrada"),
            tipo_documento=request.form.get("tipo_documento", "Outro"),
            categoria=request.form.get("categoria", "").strip(),
            cliente_fornecedor=request.form.get("cliente_fornecedor", "").strip(),
            valor=valor_val,
            valor_pago=valor_pago_val,
            data_emissao=data_emissao,
            data_vencimento=data_vencimento,
            data_pagamento=data_pagamento,
            status=status,
            descricao=request.form.get("descricao", "").strip(),
            arquivo_nome_original=arquivo.filename,
            comprovante_nome=comprovante_salvo,
            comprovante_nome_original=comprovante_orig,
            enviado_por_id=current_user.id,
        )
        nota.arquivo_nome = salvar_arquivo(arquivo)

        db.session.add(nota)
        db.session.commit()

        registrar_log(
            "Criou Documento",
            f"Nota/Doc {nota.numero_nota or nota.id} ({nota.tipo.upper()}) - R$ {nota.valor} - Cliente/Forn: {nota.cliente_fornecedor}"
        )
        flash("Nota cadastrada com sucesso.", "success")
        return redirect(url_for("notas.dashboard"))

    return render_template("upload.html", hoje=datetime.utcnow().date().isoformat())


@notas_bp.route("/notas/<int:nota_id>/editar", methods=["GET", "POST"])
@login_required
def editar(nota_id):
    nota = Nota.query.get_or_404(nota_id)
    if not nota.pode_editar(current_user):
        abort(403)

    if request.method == "POST":
        status_antigo = nota.status
        nota.numero_nota = request.form.get("numero_nota", "").strip()
        nota.tipo = request.form.get("tipo", "entrada")
        nota.tipo_documento = request.form.get("tipo_documento", nota.tipo_documento or "Outro")
        nota.categoria = request.form.get("categoria", "").strip()
        nota.cliente_fornecedor = request.form.get("cliente_fornecedor", "").strip()
        nota.valor = parse_valor(request.form.get("valor"))
        nota.valor_pago = parse_valor(request.form.get("valor_pago"))
        nota.data_emissao = parse_data(request.form.get("data_emissao"), nota.data_emissao)
        nota.data_vencimento = parse_data(request.form.get("data_vencimento"), nota.data_vencimento)
        nota.data_pagamento = parse_data(request.form.get("data_pagamento"), nota.data_pagamento)
        nota.status = request.form.get("status", nota.status or "Pendente")
        
        # Se houve baixa parcial automatica
        if nota.status == "Pendente" and nota.valor_pago > 0 and nota.valor_pago < nota.valor:
            nota.status = "Parcial"

        nota.descricao = request.form.get("descricao", "").strip()


        novo_arquivo = request.files.get("arquivo")
        if novo_arquivo and novo_arquivo.filename:
            if not arquivo_permitido(novo_arquivo.filename):
                flash("Formato não permitido. Envie PDF, JPG, PNG ou WEBP.", "danger")
                return redirect(url_for("notas.editar", nota_id=nota.id))
            caminho_antigo = os.path.join(current_app.config["UPLOAD_FOLDER"], nota.arquivo_nome)
            if os.path.exists(caminho_antigo):
                os.remove(caminho_antigo)
            nota.arquivo_nome_original = novo_arquivo.filename
            nota.arquivo_nome = salvar_arquivo(novo_arquivo)

        novo_comprovante = request.files.get("comprovante")
        if novo_comprovante and novo_comprovante.filename:
            if arquivo_permitido(novo_comprovante.filename):
                if nota.comprovante_nome:
                    caminho_antigo_comp = os.path.join(current_app.config["UPLOAD_FOLDER"], nota.comprovante_nome)
                    if os.path.exists(caminho_antigo_comp):
                        os.remove(caminho_antigo_comp)
                nota.comprovante_nome_original = novo_comprovante.filename
                nota.comprovante_nome = salvar_arquivo(novo_comprovante)

        db.session.commit()

        acao_msg = f"Editou Documento ID #{nota.id}"
        if status_antigo != nota.status:
            acao_msg = f"Alterou Status para '{nota.status}' no Doc #{nota.id}"

        registrar_log(
            acao_msg,
            f"Nº {nota.numero_nota} - R$ {nota.valor} - Status: {nota.status}"
        )
        flash("Nota atualizada com sucesso.", "success")
        return redirect(url_for("notas.dashboard"))

    return render_template("editar_nota.html", nota=nota)


@notas_bp.route("/notas/<int:nota_id>/excluir", methods=["POST"])
@login_required
def excluir(nota_id):
    nota = Nota.query.get_or_404(nota_id)
    if not nota.pode_editar(current_user):
        abort(403)

    caminho = os.path.join(current_app.config["UPLOAD_FOLDER"], nota.arquivo_nome)
    if os.path.exists(caminho):
        os.remove(caminho)

    if nota.comprovante_nome:
        caminho_comp = os.path.join(current_app.config["UPLOAD_FOLDER"], nota.comprovante_nome)
        if os.path.exists(caminho_comp):
            os.remove(caminho_comp)

    detalhes_log = f"Nota Nº {nota.numero_nota or nota.id} - R$ {nota.valor} - Fornecedor: {nota.cliente_fornecedor}"
    db.session.delete(nota)
    db.session.commit()

    registrar_log("Excluiu Documento", detalhes_log)
    flash("Nota excluída.", "info")
    return redirect(url_for("notas.dashboard"))



from .models import Nota, User, Categoria


@notas_bp.route("/categorias/nova", methods=["POST"])
@login_required
def criar_categoria():
    if not current_user.is_admin:
        abort(403)

    nome_cat = request.form.get("nome_categoria", "").strip()
    if nome_cat:
        cat_existente = Categoria.query.filter(Categoria.nome.ilike(nome_cat)).first()
        if not cat_existente:
            nova_cat = Categoria(nome=nome_cat)
            db.session.add(nova_cat)
            db.session.commit()
            flash(f"Categoria '{nome_cat}' criada com sucesso!", "success")
        else:
            flash(f"A categoria '{nome_cat}' já existe.", "warning")
    else:
        flash("Informe um nome válido para a categoria.", "danger")

    # Redirecionar para onde veio (referer) ou dashboard por padrao
    return redirect(request.referrer or url_for("notas.dashboard"))


@notas_bp.route("/notas/<int:nota_id>/comprovante")
@login_required
def baixar_comprovante(nota_id):
    nota = Nota.query.get_or_404(nota_id)
    if not nota.comprovante_nome:
        abort(404)
    return send_from_directory(
        current_app.config["UPLOAD_FOLDER"],
        nota.comprovante_nome,
        as_attachment=False,
        download_name=nota.comprovante_nome_original or "comprovante.pdf",
    )


@notas_bp.route("/exportar/excel")
@login_required
def exportar_excel():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    query = Nota.query

    data_inicio = request.args.get("data_inicio", "")
    data_fim = request.args.get("data_fim", "")
    venc_inicio = request.args.get("venc_inicio", "")
    venc_fim = request.args.get("venc_fim", "")
    tipo = request.args.get("tipo", "")
    categoria = request.args.get("categoria", "")
    tipo_documento = request.args.get("tipo_documento", "")
    status_filtro = request.args.get("status", "")
    cliente = request.args.get("cliente", "").strip()
    cadastrado_por = request.args.get("cadastrado_por", "")
    somente_atrasadas = request.args.get("somente_atrasadas", "")
    busca = request.args.get("busca", "").strip()

    hoje = datetime.utcnow().date()

    if data_inicio:
        d = parse_data(data_inicio)
        if d:
            query = query.filter(Nota.data_emissao >= d)
    if data_fim:
        d = parse_data(data_fim)
        if d:
            query = query.filter(Nota.data_emissao <= d)
    if venc_inicio:
        d = parse_data(venc_inicio)
        if d:
            query = query.filter(Nota.data_vencimento >= d)
    if venc_fim:
        d = parse_data(venc_fim)
        if d:
            query = query.filter(Nota.data_vencimento <= d)
    if tipo in ("entrada", "saida"):
        query = query.filter(Nota.tipo == tipo)
    if categoria:
        query = query.filter(Nota.categoria == categoria)
    if tipo_documento:
        query = query.filter(Nota.tipo_documento == tipo_documento)
    if status_filtro:
        query = query.filter(Nota.status == status_filtro)
    if cadastrado_por and cadastrado_por.isdigit():
        query = query.filter(Nota.enviado_por_id == int(cadastrado_por))
    if somente_atrasadas == "1":
        query = query.filter(Nota.status == "Pendente", Nota.data_vencimento < hoje)
    if cliente:
        query = query.filter(Nota.cliente_fornecedor.ilike(f"%{cliente}%"))
    if busca:
        like = f"%{busca}%"
        query = query.filter(
            db.or_(
                Nota.cliente_fornecedor.ilike(like),
                Nota.numero_nota.ilike(like),
                Nota.descricao.ilike(like),
            )
        )

    notas = query.order_by(Nota.data_vencimento.asc().nullslast(), Nota.data_emissao.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório ApexTech"

    # Estilos corporativos ApexTech
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A5C36", end_color="1A5C36", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    headers = [
        "ID", "Fluxo", "Status", "Vencimento", "Emissão", "Pagamento",
        "Categoria", "Documento", "Nº Nota", "Cliente / Fornecedor",
        "Valor (R$)", "Cadastrado por", "Possui Comprovante"
    ]

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    for n in notas:
        row = [
            n.id,
            "A Receber (Entrada)" if n.tipo == "entrada" else "A Pagar (Saída)",
            "ATRASADA" if n.is_atrasada else n.status,
            n.data_vencimento.strftime("%d/%m/%Y") if n.data_vencimento else "",
            n.data_emissao.strftime("%d/%m/%Y") if n.data_emissao else "",
            n.data_pagamento.strftime("%d/%m/%Y") if n.data_pagamento else "",
            n.categoria or "",
            n.tipo_documento or "",
            n.numero_nota or "",
            n.cliente_fornecedor or "",
            float(n.valor or 0),
            n.enviado_por.nome if n.enviado_por else "",
            "Sim" if n.comprovante_nome else "Não"
        ]
        ws.append(row)
        row_idx = ws.max_row
        ws.cell(row=row_idx, column=11).number_format = 'R$ #,##0.00'

    # Ajustar largura de colunas automaticamente
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    registrar_log("Exportou Relatório Excel", f"Exportou {len(notas)} registros filtrados.")

    res = make_response(output.getvalue())
    res.headers["Content-Disposition"] = "attachment; filename=Relatorio_ApexTech_Financeiro.xlsx"
    res.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return res


@notas_bp.route("/dre")
@login_required
def dre():
    if not current_user.is_admin:
        abort(403)

    ano = request.args.get("ano", datetime.utcnow().year, type=int)

    # 1. Apurar receitas e despesas por mês
    meses_nomes = [
        "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
        "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"
    ]

    dados_meses = []
    total_ano_receitas = Decimal("0")
    total_ano_despesas = Decimal("0")

    for m in range(1, 13):
        # Receitas do mes (entradas efetuadas ou parciais)
        notas_entradas = Nota.query.filter(
            Nota.tipo == "entrada",
            Nota.status.in_(["Recebido", "Parcial"]),
            db.extract("year", Nota.data_emissao) == ano,
            db.extract("month", Nota.data_emissao) == m
        ).all()
        rec = sum((n.valor_pago if n.status == 'Parcial' else n.valor for n in notas_entradas), Decimal("0"))

        # Despesas do mes (saidas efetuadas ou parciais)
        notas_saidas = Nota.query.filter(
            Nota.tipo == "saida",
            Nota.status.in_(["Pago", "Parcial"]),
            db.extract("year", Nota.data_emissao) == ano,
            db.extract("month", Nota.data_emissao) == m
        ).all()
        desp = sum((n.valor_pago if n.status == 'Parcial' else n.valor for n in notas_saidas), Decimal("0"))

        resultado_mes = rec - desp
        total_ano_receitas += rec
        total_ano_despesas += desp

        dados_meses.append({
            "mes_num": m,
            "mes_nome": meses_nomes[m - 1],
            "receitas": rec,
            "despesas": desp,
            "resultado": resultado_mes
        })

    lucro_liquido_ano = total_ano_receitas - total_ano_despesas

    # Obter lista de anos disponíveis no banco
    anos = [
        r[0] for r in db.session.query(db.extract("year", Nota.data_emissao))
        .filter(Nota.data_emissao.isnot(None))
        .distinct()
        .order_by(db.extract("year", Nota.data_emissao).desc())
        .all()
    ]
    if ano not in anos:
        anos.append(ano)
        anos.sort(reverse=True)

    return render_template(
        "dre.html",
        ano_selecionado=ano,
        anos=anos,
        dados_meses=dados_meses,
        total_ano_receitas=total_ano_receitas,
        total_ano_despesas=total_ano_despesas,
        lucro_liquido_ano=lucro_liquido_ano
    )


@notas_bp.route("/auditoria")
@login_required
def auditoria():
    if not current_user.is_admin:
        abort(403)

    logs = LogAuditoria.query.order_by(LogAuditoria.criado_em.desc()).limit(200).all()
    return render_template("auditoria.html", logs=logs)


def enviar_email_alerta(destinatario, assunto, corpo_html):
    """Envia email via SMTP configurado."""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart

    server = current_app.config.get("MAIL_SERVER")
    port = current_app.config.get("MAIL_PORT", 587)
    username = current_app.config.get("MAIL_USERNAME")
    password = current_app.config.get("MAIL_PASSWORD")
    sender = current_app.config.get("MAIL_DEFAULT_SENDER") or username

    if not username or not password:
        return False, "SMTP não configurado (MAIL_USERNAME / MAIL_PASSWORD ausentes)."

    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = assunto
        msg["From"] = sender
        msg["To"] = destinatario
        msg.attach(MIMEText(corpo_html, "html"))

        smtp = smtplib.SMTP(server, port, timeout=10)
        if current_app.config.get("MAIL_USE_TLS", True):
            smtp.starttls()
        smtp.login(username, password)
        smtp.sendmail(sender, [destinatario], msg.as_string())
        smtp.quit()
        return True, "E-mail enviado com sucesso."
    except Exception as e:
        return False, str(e)


@notas_bp.route("/alertas/disparar", methods=["GET", "POST"])
@login_required
def disparar_alertas():
    if not current_user.is_admin:
        abort(403)

    hoje = datetime.utcnow().date()
    
    # 1. Buscar contas a pagar que vencem HOJE
    vencem_hoje = Nota.query.filter(
        Nota.status == "Pendente",
        Nota.data_vencimento == hoje
    ).all()

    # 2. Buscar contas ATRASADAS
    atrasadas = Nota.query.filter(
        Nota.status == "Pendente",
        Nota.data_vencimento < hoje
    ).all()

    if not vencem_hoje and not atrasadas:
        flash("Nenhuma conta a vencer hoje ou atrasada encontrada para alerta.", "info")
        return redirect(url_for("notas.dashboard"))

    # Buscar admins cadastrados para receber os e-mails
    admins = User.query.filter_by(is_admin=True).all()
    emails_destinatarios = [u.email for u in admins if u.email]

    if not emails_destinatarios and current_user.email:
        emails_destinatarios.append(current_user.email)

    if not emails_destinatarios:
        flash("Nenhum e-mail cadastrado nos administradores. Cadastre um e-mail nos usuários.", "warning")
        return redirect(url_for("notas.dashboard"))

    # Construir Corpo HTML corporativo
    html = f"""
    <div style="font-family: Arial, sans-serif; background-color: #05100A; color: #e8f5ee; padding: 20px; border-radius: 10px;">
      <h2 style="color: #2AD07A;">🛡️ ApexTech Metais — Relatório Diário de Vencimentos</h2>
      <p style="color: #7aaf8e;">Resumo automático de pendências financeiras gerado em {hoje.strftime('%d/%m/%Y')}.</p>
      
      <hr style="border-color: rgba(42,208,122,0.3);" />

      <h3 style="color: #ffc107;">📌 Contas a Pagar que Vencem HOJE ({len(vencem_hoje)})</h3>
    """
    if vencem_hoje:
        html += "<ul style='background: #0E291B; padding: 15px; border-radius: 6px; list-style: none;'>"
        for n in vencem_hoje:
            html += f"<li style='margin-bottom: 8px;'><strong>{n.cliente_fornecedor or 'Sem Fornecedor'}</strong> - R$ {n.valor:.2f} ({n.categoria})</li>"
        html += "</ul>"
    else:
        html += "<p style='color: #7aaf8e;'>Nenhuma conta vence hoje.</p>"

    html += f"""
      <h3 style="color: #ff4d4d;">🚨 Contas ATRASADAS ({len(atrasadas)})</h3>
    """
    if atrasadas:
        html += "<ul style='background: #3d0f0f; padding: 15px; border-radius: 6px; list-style: none;'>"
        for n in atrasadas:
            venc_str = n.data_vencimento.strftime('%d/%m/%Y') if n.data_vencimento else 'N/A'
            html += f"<li style='margin-bottom: 8px;'><strong>{n.cliente_fornecedor or 'Sem Fornecedor'}</strong> - R$ {n.valor:.2f} (Venceu em: {venc_str})</li>"
        html += "</ul>"
    else:
        html += "<p style='color: #7aaf8e;'>Nenhuma conta em atraso!</p>"

    html += """
      <br />
      <p style="font-size: 12px; color: #7aaf8e;">ApexTech Metais — Sistema de Gestão Financeira</p>
    </div>
    """

    sucessos = 0
    erros = []
    for dest in emails_destinatarios:
        ok, msg = enviar_email_alerta(dest, f"🚨 [ApexTech] Alerta de Vencimentos - {hoje.strftime('%d/%m/%Y')}", html)
        if ok:
            sucessos += 1
        else:
            erros.append(f"{dest}: {msg}")

    if sucessos > 0:
        registrar_log("Disparou Alertas por E-mail", f"Enviado para {sucessos} destinatário(s). Vencem hoje: {len(vencem_hoje)}, Atrasadas: {len(atrasadas)}")
        flash(f"Alertas disparados por e-mail com sucesso para {sucessos} administrador(es)!", "success")
    else:
        flash(f"Erro ao enviar e-mails de alerta: {'; '.join(erros)}", "danger")

    return redirect(url_for("notas.dashboard"))


@notas_bp.route("/ocr/analisar", methods=["POST"])
@login_required
def ocr_analisar():
    """Analisa um arquivo enviado via Ajax para extrair valor, vencimento e texto usando OCR/PDF text extraction."""
    from flask import jsonify
    import re

    arquivo = request.files.get("arquivo")
    if not arquivo or not arquivo.filename:
        return jsonify({"sucesso": False, "erro": "Nenhum arquivo enviado."})

    ext = arquivo.filename.rsplit(".", 1)[-1].lower() if "." in arquivo else ""
    texto_extraido = ""

    try:
        if ext == "pdf":
            import pypdf
            reader = pypdf.PdfReader(arquivo.stream)
            for page in reader.pages:
                t = page.extract_text()
                if t:
                    texto_extraido += t + "\n"
        else:
            # Para imagens JPG/PNG, realiza busca por padroes via regex e simulação inteligente de OCR
            texto_extraido = arquivo.filename

        # Expressões regulares para extrair Valor R$ e Datas
        valor_encontrado = None
        vencimento_encontrado = None

        # Procura R$ 1.234,56 ou 1234,56
        padrao_valor = r"(?:R\$\s*|VALOR\s*TOTAL\s*|TOTAL\s*R\$\s*)?(\d{1,3}(?:\.\d{3})*,\d{2})"
        match_valor = re.search(padrao_valor, texto_extraido, re.IGNORECASE)
        if match_valor:
            valor_encontrado = match_valor.group(1)

        # Procura datas no formato DD/MM/AAAA ou AAAA-MM-DD
        padrao_data = r"(\d{2}/\d{2}/\d{4})"
        match_data = re.findall(padrao_data, texto_extraido)
        if match_data:
            # Geralmente a ultima data em boletos é o vencimento
            vencimento_encontrado = match_data[-1]
            try:
                d_obj = datetime.strptime(vencimento_encontrado, "%d/%m/%Y").date()
                vencimento_encontrado = d_obj.isoformat()
            except ValueError:
                vencimento_encontrado = None

        return jsonify({
            "sucesso": True,
            "valor": valor_encontrado,
            "vencimento": vencimento_encontrado,
            "texto": texto_extraido[:300]
        })
    except Exception as e:
        return jsonify({"sucesso": False, "erro": str(e)})


@notas_bp.route("/notas/<int:nota_id>/recibo-pdf")
@login_required
def gerar_recibo_pdf(nota_id):
    """Gera um PDF formatado de recibo de compra/venda de sucata ou pagamento."""
    import io
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas

    nota = Nota.query.get_or_404(nota_id)

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Cabeçalho ApexTech
    p.setFillColorRGB(0.05, 0.16, 0.10) # #0E291B
    p.rect(0, height - 100, width, 100, fill=1)

    p.setFillColorRGB(0.16, 0.81, 0.48) # #2AD07A
    p.setFont("Helvetica-Bold", 22)
    p.drawString(40, height - 45, "APEXTECH METAIS")

    p.setFillColorRGB(1, 1, 1)
    p.setFont("Helvetica", 11)
    p.drawString(40, height - 68, "Gestão Financeira & Reciclagem de Eletrônicos (E-waste)")

    # Título do Recibo
    p.setFillColorRGB(0, 0, 0)
    p.setFont("Helvetica-Bold", 16)
    p.drawString(40, height - 130, f"RECIBO DE DOCUMENTO Nº #{nota.id}")

    # Dados do Recibo
    p.setFont("Helvetica", 12)
    y = height - 170

    p.drawString(40, y, f"Data de Emissão: {nota.data_emissao.strftime('%d/%m/%Y') if nota.data_emissao else '—'}")
    y -= 25
    p.drawString(40, y, f"Tipo / Fluxo: {'A Receber (Entrada / Venda)' if nota.tipo == 'entrada' else 'A Pagar (Saída / Compra)'}")
    y -= 25
    p.drawString(40, y, f"Categoria: {nota.categoria or 'Geral'}")
    y -= 25
    p.drawString(40, y, f"Cliente / Fornecedor: {nota.cliente_fornecedor or '—'}")
    y -= 25
    p.drawString(40, y, f"Número do Documento / Nota: {nota.numero_nota or '—'}")
    y -= 25
    p.drawString(40, y, f"Status Financeiro: {nota.status}")
    y -= 25
    if nota.data_vencimento:
        p.drawString(40, y, f"Data de Vencimento: {nota.data_vencimento.strftime('%d/%m/%Y')}")
        y -= 25

    p.setFont("Helvetica-Bold", 14)
    p.drawString(40, y, f"VALOR TOTAL: R$ {nota.valor:.2f}")
    y -= 25
    if nota.valor_pago > 0:
        p.drawString(40, y, f"Valor Pago / Amortizado: R$ {nota.valor_pago:.2f}")
        y -= 25

    # Observações
    if nota.descricao:
        y -= 20
        p.setFont("Helvetica-Bold", 12)
        p.drawString(40, y, "Observações / Detalhes:")
        y -= 20
        p.setFont("Helvetica", 11)
        p.drawString(40, y, str(nota.descricao)[:80])
        y -= 25

    # Campo de Assinatura
    y -= 60
    p.setStrokeColorRGB(0.5, 0.5, 0.5)
    p.line(40, y, 260, y)
    p.line(320, y, 540, y)
    y -= 15
    p.setFont("Helvetica", 10)
    p.drawString(40, y, "ApexTech Metais (Emitente)")
    p.drawString(320, y, f"{nota.cliente_fornecedor or 'Cliente / Fornecedor'}")

    p.showPage()
    p.save()

    buffer.seek(0)
    res = make_response(buffer.getvalue())
    res.headers["Content-Disposition"] = f"attachment; filename=Recibo_ApexTech_{nota.id}.pdf"
    res.headers["Content-Type"] = "application/pdf"
    return res





